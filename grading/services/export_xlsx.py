"""
Export service for generating Excel files.
"""
from io import BytesIO
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


class ExcelExportService:
    """
    Service for exporting grading results to Excel.
    """
    
    # Column headers
    HEADERS = [
        ('Öğrenci No', 15),
        ('Ad Soyad', 25),
        ('Kitapçık', 10),
        ('Doğru', 10),
        ('Yanlış', 10),
        ('Boş', 10),
        ('Geçersiz', 10),
        ('Net', 10),
        ('Puan', 10),
    ]
    
    def __init__(self):
        """Initialize export service."""
        self.header_font = Font(bold=True)
        self.header_fill = PatternFill(start_color='DAEEF3', end_color='DAEEF3', fill_type='solid')
        self.header_alignment = Alignment(horizontal='center', vertical='center')
        self.highlight_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    
    def export_results(self, results: List[dict], penalty_ratio: int = 0, points_per_question: float = 1.0) -> BytesIO:
        """
        Export results to Excel file.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = 'Sonuçlar'
        
        # Write settings
        self._write_settings(ws, penalty_ratio, points_per_question)
        
        # Write headers
        self._write_headers(ws)
        
        # Write data rows
        for row_num, result in enumerate(results, start=2):
            self._write_result_row(ws, row_num, result)
        
        # Apply column widths
        self._set_column_widths(ws)
        ws.auto_filter.ref = f"A1:{get_column_letter(len(self.HEADERS))}1"
        ws.freeze_panes = 'A2'
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    def export_with_details(self, results: List[dict], question_count: int, penalty_ratio: int = 0, points_per_question: float = 1.0, stats: dict = None) -> BytesIO:
        """
        Export results with detailed analysis and statistics.
        """
        wb = Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = 'Özet Sonuçlar'
        
        # Write settings
        self._write_settings(ws_summary, penalty_ratio, points_per_question)
        
        # Write headers
        self._write_headers(ws_summary)
        
        for row_num, result in enumerate(results, start=2):
            self._write_result_row(ws_summary, row_num, result)
        
        self._set_column_widths(ws_summary)
        ws_summary.auto_filter.ref = f"A1:{get_column_letter(len(self.HEADERS))}1"
        ws_summary.freeze_panes = 'A2'
        
        # Detail sheet
        ws_detail = wb.create_sheet('Soru Detayları')
        self._write_detail_sheet(ws_detail, results, question_count)
        
        # Statistics sheet
        if stats:
            ws_stats = wb.create_sheet('İstatistikler')
            self._write_statistics_sheet(ws_stats, stats)
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output

    def _write_settings(self, ws, penalty_ratio: int, points_per_question: float):
        """Write calculation settings to the worksheet (hidden or off to the side)."""
        # We'll place them at K1:L2
        ws.cell(row=1, column=11, value='Yanlış Oranı:').font = self.header_font
        ws.cell(row=1, column=12, value=penalty_ratio).fill = self.highlight_fill
        
        ws.cell(row=2, column=11, value='Soru Puanı:').font = self.header_font
        ws.cell(row=2, column=12, value=points_per_question).fill = self.highlight_fill
        
        ws.column_dimensions['K'].width = 15
        ws.column_dimensions['L'].width = 10

    def _write_result_row(self, ws, row_num, result):
        """Write a single student result row with formulas."""
        ws.cell(row=row_num, column=1, value=result.get('student_no', ''))
        ws.cell(row=row_num, column=2, value=result.get('student_name', ''))
        ws.cell(row=row_num, column=3, value=result.get('booklet', ''))
        ws.cell(row=row_num, column=4, value=result.get('correct_count', 0))
        ws.cell(row=row_num, column=5, value=result.get('wrong_count', 0))
        ws.cell(row=row_num, column=6, value=result.get('blank_count', 0))
        ws.cell(row=row_num, column=7, value=result.get('invalid_count', 0))
        
        # Net Formula (Column H): Correct - (Wrong / Ratio)
        # Handle ratio 0 case (though usually it's 3 or 4)
        ws.cell(row=row_num, column=8, value=f'=D{row_num}-IF($L$1>0, E{row_num}/$L$1, 0)')
        
        # Score Formula (Column I): Net * Points
        ws.cell(row=row_num, column=9, value=f'=H{row_num}*$L$2')

    def _write_detail_sheet(self, ws_detail, results, question_count):
        # Headers for detail sheet
        detail_headers = ['Öğrenci No', 'Ad Soyad'] + [f'S{i+1}' for i in range(question_count)]
        for col, header in enumerate(detail_headers, start=1):
            cell = ws_detail.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
        
        # Data rows
        for row_num, result in enumerate(results, start=2):
            ws_detail.cell(row=row_num, column=1, value=result.get('student_no', ''))
            ws_detail.cell(row=row_num, column=2, value=result.get('student_name', ''))
            
            detailed = result.get('detailed_results', '')
            for q_idx, status in enumerate(detailed):
                cell = ws_detail.cell(row=row_num, column=3 + q_idx, value=status)
                # Color coding
                if status == 'D':
                    cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                elif status == 'Y':
                    cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                elif status == 'B':
                    cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                elif status == 'G':
                    cell.fill = PatternFill(start_color='A6A6A6', end_color='A6A6A6', fill_type='solid')
        
        # Set column widths
        ws_detail.column_dimensions['A'].width = 15
        ws_detail.column_dimensions['B'].width = 25
        for i in range(question_count):
            ws_detail.column_dimensions[get_column_letter(3 + i)].width = 5
        
        ws_detail.freeze_panes = 'C2'

    def _write_statistics_sheet(self, ws, stats):
        """Write statistics and item analysis to a sheet."""
        # General stats
        row = 1
        ws.cell(row=row, column=1, value='Genel İstatistikler').font = Font(bold=True, size=14)
        row += 2
        
        general_data = [
            ('Öğrenci Sayısı', stats.get('student_count')),
            ('Soru Sayısı', stats.get('question_count')),
            ('Ortalama', stats.get('mean')),
            ('Medyan', stats.get('median')),
            ('Standart Sapma', stats.get('std_dev')),
            ('Değişim Katsayısı', stats.get('variation_coeff')),
        ]
        
        for label, value in general_data:
            ws.cell(row=row, column=1, value=label).font = self.header_font
            ws.cell(row=row, column=2, value=value)
            row += 1
            
        row += 1
        ws.cell(row=row, column=1, value='Dağılım Yorumu').font = self.header_font
        ws.cell(row=row, column=2, value=stats.get('dist_interpretation'))
        row += 1
        ws.cell(row=row, column=1, value='Grup Yapısı').font = self.header_font
        ws.cell(row=row, column=2, value=stats.get('group_structure'))
        
        row += 3
        ws.cell(row=row, column=1, value='Madde Analizi').font = Font(bold=True, size=14)
        row += 2
        
        # Item analysis headers
        item_headers = ['Soru No', 'Cevap', 'Güçlük (p)', 'Ayırt Edicilik (r)', 'Yorum (p)', 'Yorum (r)']
        for col, header in enumerate(item_headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            
        analysis_row = row + 1
        for item in stats.get('item_analysis', []):
            ws.cell(row=analysis_row, column=1, value=item['question_number'])
            ws.cell(row=analysis_row, column=2, value=item['correct_answer'])
            ws.cell(row=analysis_row, column=3, value=round(item['p'], 2))
            ws.cell(row=analysis_row, column=4, value=round(item['r'], 2))
            ws.cell(row=analysis_row, column=5, value=item['difficulty'])
            ws.cell(row=analysis_row, column=6, value=item['discrimination'])
            analysis_row += 1
            
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20

    def _write_headers(self, ws):
        """Write headers to worksheet."""
        for col, (header, width) in enumerate(self.HEADERS, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
    
    def _set_column_widths(self, ws):
        """Set column widths."""
        for col, (header, width) in enumerate(self.HEADERS, start=1):
            ws.column_dimensions[get_column_letter(col)].width = width
    
    def _write_headers(self, ws):
        """Write headers to worksheet."""
        for col, (header, width) in enumerate(self.HEADERS, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
    
    def _set_column_widths(self, ws):
        """Set column widths."""
        for col, (header, width) in enumerate(self.HEADERS, start=1):
            ws.column_dimensions[get_column_letter(col)].width = width
